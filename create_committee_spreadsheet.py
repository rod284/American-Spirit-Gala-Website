from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
wb.remove(wb.active)

# SHEET 1: Committee Members
ws1 = wb.create_sheet('Committee Members', 0)
headers = ['First Name', 'Last Name', 'Role(s)', 'Phone', 'Email', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

data = [
    ['Rod', 'Dennis', 'Theme, Leadership', '(480) 695-0733', 'rod@roddennis.com', ''],
    ['Jeff', 'Little', 'Corporate Sponsors', '(480) 686-2992', 'littlejeff@gmail.com', ''],
    ['Valentine', 'Rhodes', 'Committee Member', '(480) 213-2360', '', ''],
    ['Dolf', 'May', 'Auction - Lead', '(602) 616-2159', '', 'Secure vendor, items, auctioneer by 8/1'],
    ['Mike', 'Yarnall', 'Run of House, Setup/Takedown', '(602) 576-0492', '', 'MC, Program, Logistics'],
    ['Coleman', 'Caldwell', 'Member Recruitment', '', '', 'Booth staffing, materials'],
    ['Brendan', 'Cassin', 'Committee Member', '(331) 643-9419', '', ''],
    ['Larry', 'Hewitt', 'Committee Member', '', '', ''],
    ['Brian', 'Marcy', 'Committee Member', '', '', ''],
    ['Chris', 'Harris', 'Corporate Sponsors Support', '', '', ''],
    ['Neil', 'Wilson', 'Tickets, Private Donors, Entertainment TBD', '', '', 'Get prior year materials from Jim'],
    ['Steve', 'Denny', 'Other Funding Events', '', '', '50/50, raffle, ring game options'],
    ['Mike', 'Mcmahon', 'Finance', '', '', ''],
    ['Seabold', '', 'Setup/Takedown Lead', '', '', 'Day-of logistics'],
    ['Jim', 'Kilroy', 'Corporate Sponsors Support', '', '', 'Prior year materials/donor list'],
    ['Jon', 'Mirmelli', 'Corporate Sponsors Support', '', '', 'Recruit Darren/Yarnell'],
    ['Zach', '', 'Private Donors, Setup/Takedown', '', '', 'Prior year info for Coleman'],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 25
ws1.column_dimensions['F'].width = 40

# SHEET 2: Committee Roles & Assignments
ws2 = wb.create_sheet('Roles & Assignments', 1)
role_data = [
    ['Category', 'Lead', 'Support', 'Responsibility', 'Deadline/Notes'],
    ['Auction', 'Dolf May', 'Neil Wilson (TBD)', 'Secure online vendor, gather auction items, secure auctioneer', '8/1 - Finalize auctioneer'],
    ['Entertainment', 'Neil Wilson', 'TBD', 'Select/contract band/DJ. JMT3 (last year) or Jeff L free/cheap option', 'Band sponsor needed'],
    ['Member Recruitment', 'Coleman Caldwell', 'Zach', 'Recruitment station materials and booth staffing', 'Get prior year info from Zach'],
    ['Corporate Sponsors', 'Jeff Little', 'Chris Harris, Jim Kilroy, Jon Mirmelli', 'Outreach, design funding levels, exceed revenue targets', 'Get prior year materials from Jim'],
    ['Tickets', 'Neil Wilson', '', 'Printing and selling tickets, distribution to donors/sponsors', 'Get tickets for August meeting'],
    ['Private Donors', 'Neil Wilson', 'Zach', 'Solicit past donors, create materials for member recruitment', 'Mike to follow up on progress'],
    ['Funding Events', 'Steve Denny', '', '50/50, ring game, raffle, auction item raffle', 'Review options with auctioneer'],
    ['Theme', 'Rod Dennis', '', 'Create theme materials for promotion (Cowboys/Western or Red/White/Blue)', 'Decide by next meeting'],
    ['Run of House', 'Mike Yarnall', '', 'Program flow, optimal event timing, MC duties', 'Program needed for distribution'],
    ['Setup/Takedown', 'Seabold', 'Zach, Mike Franks', 'Organize all logistics for day-of setup and breakdown', 'Menu planning needed'],
    ['Finance', 'Mike Mcmahon', '', 'Budget management and revenue tracking', ''],
]

for row_idx, row_data in enumerate(role_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        if row_idx == 1:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 40
ws2.column_dimensions['E'].width = 30

# SHEET 3: Action Items
ws3 = wb.create_sheet('Action Items', 2)
action_headers = ['Action Item', 'Owner', 'Deadline', 'Status', 'Notes']
for col, header in enumerate(action_headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

action_data = [
    ['Finalize auctioneer', 'Dolf May', '8/1/2026', 'Pending', 'Two proposals by mid-July'],
    ['Finalize Perry Consulting contract', 'Dolf May', 'TBD', 'Pending', ''],
    ['Select band/DJ', 'Neil Wilson', 'TBD', 'In Progress', 'Evaluating JMT3 and Jeff L options'],
    ['Find band sponsor', 'Jeff Little', 'TBD', 'Pending', ''],
    ['Gather auction items', 'Dolf May', 'TBD', 'In Progress', 'Based on prior year items'],
    ['Decide on theme', 'Rod Dennis', 'Next Meeting', 'Pending', 'Cowboys/Western or Red/White/Blue'],
    ['Get prior year materials', 'Jim Kilroy', 'Before Next Meeting', 'Pending', 'Corporate sponsor list and program structure'],
    ['Get tickets for August', 'Neil Wilson', '8/31/2026', 'Pending', 'Printed tickets needed'],
    ['Email blast - Prior donors', 'Neil Wilson', 'TBD', 'Pending', 'Coordinate with Troy'],
]

for row_idx, row_data in enumerate(action_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 30

wb.save('C:\\Users\\rodde\\OneDrive\\Documentos\\AllHandsOnDeck\\COMMITTEE_TRACKING.xlsx')
print("Committee tracking spreadsheet created!")
